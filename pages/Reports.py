import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# =========================================================
# PAGE CONFIG
# =========================================================
st.set_page_config(
    page_title="MoneyMate Reports",
    page_icon="📑",
    layout="wide"
)


# =========================================================
# SUPABASE
# =========================================================
from utils.supabase_client import supabase


# =========================================================
# RESTORE SESSION
# =========================================================
if (
    "access_token" in st.session_state
    and "refresh_token" in st.session_state
):
    try:
        supabase.auth.set_session(
            st.session_state["access_token"],
            st.session_state["refresh_token"]
        )
    except Exception:
        pass


# =========================================================
# LOGIN CHECK
# =========================================================
if "user" not in st.session_state or not st.session_state["user"]:
    st.warning("Please log in first.")
    st.stop()

user = st.session_state["user"]


# =========================================================
# USER SETTINGS
# =========================================================
try:
    metadata = user.user_metadata or {}
except Exception:
    metadata = {}


display_name = metadata.get(
    "display_name",
    ""
)

currency_name = metadata.get(
    "currency",
    "₹ INR"
)

custom_categories = metadata.get(
    "custom_categories",
    []
)

custom_accounts = metadata.get(
    "custom_accounts",
    []
)


# =========================================================
# CURRENCY
# =========================================================
CURRENCY_SYMBOLS = {
    "₹ INR": "₹",
    "$ USD": "$",
    "€ EUR": "€",
    "£ GBP": "£",
    "¥ JPY": "¥",
    "₩ KRW": "₩",
    "AED": "AED ",
    "SGD": "SGD "
}

currency_symbol = CURRENCY_SYMBOLS.get(
    currency_name,
    "₹"
)


def money(value):
    try:
        return f"{currency_symbol}{float(value):,.2f}"
    except Exception:
        return f"{currency_symbol}0.00"


# =========================================================
# HEADER
# =========================================================
if display_name.strip():
    st.title(
        f"📑 {display_name.strip()}'s Reports"
    )
else:
    st.title("📑 MoneyMate Reports")

st.caption(
    "Income, expenses, budgets and financial reports"
)


# =========================================================
# LOAD TRANSACTIONS
# =========================================================
try:
    transaction_response = (
        supabase
        .table("transactions")
        .select("*")
        .eq("user_id", user.id)
        .order("date", desc=True)
        .execute()
    )

    transaction_data = (
        transaction_response.data or []
    )

    df = pd.DataFrame(
        transaction_data
    )

except Exception as e:
    st.error(
        f"Unable to load transactions: {e}"
    )
    st.stop()


# =========================================================
# EMPTY DATAFRAME STRUCTURE
# =========================================================
if df.empty:
    df = pd.DataFrame(
        columns=[
            "id",
            "user_id",
            "date",
            "type",
            "amount",
            "category",
            "account",
            "note"
        ]
    )


# =========================================================
# REQUIRED COLUMNS
# =========================================================
for column in [
    "date",
    "type",
    "amount",
    "category",
    "account",
    "note"
]:
    if column not in df.columns:
        df[column] = ""


# =========================================================
# CLEAN TRANSACTIONS
# =========================================================
df["date"] = pd.to_datetime(
    df["date"],
    errors="coerce"
)

df["amount"] = (
    df["amount"]
    .astype(str)
    .str.replace("₹", "", regex=False)
    .str.replace(",", "", regex=False)
    .str.strip()
)

df["amount"] = pd.to_numeric(
    df["amount"],
    errors="coerce"
).fillna(0.0)

df["type"] = (
    df["type"]
    .fillna("")
    .astype(str)
    .str.strip()
    .str.lower()
)

df["category"] = (
    df["category"]
    .fillna("")
    .astype(str)
    .str.strip()
)

df["account"] = (
    df["account"]
    .fillna("")
    .astype(str)
    .str.strip()
)

df["note"] = (
    df["note"]
    .fillna("")
    .astype(str)
    .str.strip()
)

df = df.dropna(
    subset=["date"]
).copy()


# =========================================================
# TITLE FILTER
# =========================================================
st.divider()
st.subheader("📅 Report Period")


filter_type = st.radio(
    "Select Report Type",
    [
        "Monthly",
        "Annual",
        "Custom Range"
    ],
    horizontal=True,
    key="report_type"
)


# =========================================================
# MONTHLY
# =========================================================
if filter_type == "Monthly":

    selected_month = st.date_input(
        "Select Month",
        value=date.today().replace(day=1),
        key="report_month"
    )

    month_start = selected_month.replace(
        day=1
    )

    if month_start.month == 12:
        period_end = month_start.replace(
            year=month_start.year + 1,
            month=1
        )
    else:
        period_end = month_start.replace(
            month=month_start.month + 1
        )

    report_df = df[
        (df["date"] >= pd.Timestamp(month_start))
        &
        (df["date"] < pd.Timestamp(period_end))
    ].copy()

    report_title = month_start.strftime(
        "%B %Y"
    )


# =========================================================
# ANNUAL
# =========================================================
elif filter_type == "Annual":

    available_years = sorted(
        df["date"].dt.year.unique(),
        reverse=True
    )

    if not available_years:
        available_years = [
            date.today().year
        ]

    selected_year = st.selectbox(
        "Select Year",
        available_years,
        key="report_year"
    )

    report_df = df[
        df["date"].dt.year == selected_year
    ].copy()

    report_title = str(
        selected_year
    )


# =========================================================
# CUSTOM RANGE
# =========================================================
else:

    col1, col2 = st.columns(2)

    with col1:
        start_date = st.date_input(
            "Start Date",
            value=date(
                date.today().year,
                1,
                1
            ),
            key="report_start"
        )

    with col2:
        end_date = st.date_input(
            "End Date",
            value=date.today(),
            key="report_end"
        )

    if start_date > end_date:
        st.error(
            "Start date cannot be after end date."
        )
        st.stop()

    report_df = df[
        (df["date"] >= pd.Timestamp(start_date))
        &
        (df["date"] < pd.Timestamp(end_date) + pd.Timedelta(days=1))
    ].copy()

    report_title = (
        f"{start_date.strftime('%d-%m-%Y')} "
        f"to "
        f"{end_date.strftime('%d-%m-%Y')}"
    )


# =========================================================
# NO DATA
# =========================================================
if report_df.empty:

    st.info(
        f"No transactions found for {report_title}."
    )

    # Still allow empty report view
    income_df = report_df.copy()
    expense_df = report_df.copy()

else:

    income_df = report_df[
        report_df["type"] == "income"
    ].copy()

    expense_df = report_df[
        report_df["type"] == "expense"
    ].copy()


# =========================================================
# TOTALS
# =========================================================
total_income = float(
    income_df["amount"].sum()
)

total_expense = float(
    expense_df["amount"].sum()
)

balance = (
    total_income -
    total_expense
)

if total_income > 0:
    savings_percent = (
        balance /
        total_income
    ) * 100
else:
    savings_percent = 0


if total_income > 0:
    expense_percent = (
        total_expense /
        total_income
    ) * 100
else:
    expense_percent = 0


# =========================================================
# KPI
# =========================================================
st.divider()

st.subheader(
    f"📊 Financial Summary — {report_title}"
)

k1, k2, k3, k4 = st.columns(4)

with k1:
    st.metric(
        "💰 Income",
        money(total_income)
    )

with k2:
    st.metric(
        "💸 Expense",
        money(total_expense)
    )

with k3:
    st.metric(
        "🏦 Balance",
        money(balance)
    )

with k4:
    st.metric(
        "📈 Savings",
        f"{savings_percent:.1f}%"
    )


# =========================================================
# BUDGET DATA
# =========================================================
budget_df = pd.DataFrame()

try:

    budget_response = (
        supabase
        .table("budgets")
        .select("*")
        .eq("user_id", user.id)
        .execute()
    )

    budget_df = pd.DataFrame(
        budget_response.data or []
    )

except Exception as e:

    st.warning(
        f"Budget data could not be loaded: {e}"
    )


# =========================================================
# CLEAN BUDGET
# =========================================================
if not budget_df.empty:

    if "amount" not in budget_df.columns:
        budget_df["amount"] = 0.0

    budget_df["amount"] = pd.to_numeric(
        budget_df["amount"],
        errors="coerce"
    ).fillna(0.0)

    if "category" not in budget_df.columns:
        budget_df["category"] = ""

    if "month" not in budget_df.columns:
        budget_df["month"] = ""


# =========================================================
# BUDGET REPORT
# =========================================================
st.divider()
st.subheader("💰 Budget vs Actual")


budget_report = pd.DataFrame()


if filter_type == "Monthly":

    current_budget = budget_df[
        budget_df["month"].astype(str)
        == str(month_start)
    ].copy() if not budget_df.empty else pd.DataFrame()

elif filter_type == "Annual":

    current_budget = budget_df[
        budget_df["month"]
        .astype(str)
        .str.startswith(str(selected_year))
    ].copy() if not budget_df.empty else pd.DataFrame()

else:

    if not budget_df.empty:

        budget_temp = budget_df.copy()

        budget_temp["month_date"] = pd.to_datetime(
            budget_temp["month"],
            errors="coerce"
        )

        current_budget = budget_temp[
            (budget_temp["month_date"] >= pd.Timestamp(start_date))
            &
            (budget_temp["month_date"] <= pd.Timestamp(end_date))
        ].copy()

    else:
        current_budget = pd.DataFrame()


# =========================================================
# BUILD BUDGET REPORT
# =========================================================
if not current_budget.empty:

    actual_by_category = (
        expense_df
        .groupby("category")["amount"]
        .sum()
        .to_dict()
    )

    budget_rows = []

    for _, row in current_budget.iterrows():

        category = str(
            row.get(
                "category",
                ""
            )
        )

        budget_amount = float(
            row.get(
                "amount",
                0
            )
        )

        actual_amount = float(
            actual_by_category.get(
                category,
                0
            )
        )

        remaining = (
            budget_amount -
            actual_amount
        )

        if budget_amount > 0:
            utilization = (
                actual_amount /
                budget_amount
            ) * 100
        else:
            utilization = 0

        if utilization > 100:
            status = "🔴 Over Budget"

        elif utilization >= 80:
            status = "🟠 Near Limit"

        else:
            status = "🟢 On Track"

        budget_rows.append({
            "Category": category,
            "Budget": budget_amount,
            "Actual": actual_amount,
            "Remaining": remaining,
            "Utilization": utilization,
            "Status": status
        })

    budget_report = pd.DataFrame(
        budget_rows
    )


if budget_report.empty:

    st.info(
        "No budget has been created for this report period."
    )

else:

    budget_total = float(
        budget_report["Budget"].sum()
    )

    budget_actual = float(
        budget_report["Actual"].sum()
    )

    budget_remaining = (
        budget_total -
        budget_actual
    )

    if budget_total > 0:
        budget_used_percent = (
            budget_actual /
            budget_total
        ) * 100
    else:
        budget_used_percent = 0


    b1, b2, b3, b4 = st.columns(4)

    with b1:
        st.metric(
            "Budget",
            money(budget_total)
        )

    with b2:
        st.metric(
            "Actual",
            money(budget_actual)
        )

    with b3:
        st.metric(
            "Remaining",
            money(budget_remaining)
        )

    with b4:
        st.metric(
            "Used",
            f"{budget_used_percent:.1f}%"
        )


    display_budget = budget_report.copy()

    display_budget["Budget"] = (
        display_budget["Budget"]
        .apply(money)
    )

    display_budget["Actual"] = (
        display_budget["Actual"]
        .apply(money)
    )

    display_budget["Remaining"] = (
        display_budget["Remaining"]
        .apply(money)
    )

    display_budget["Utilization"] = (
        display_budget["Utilization"]
        .apply(
            lambda x: f"{x:.1f}%"
        )
    )

    st.dataframe(
        display_budget[
            [
                "Category",
                "Budget",
                "Actual",
                "Remaining",
                "Utilization",
                "Status"
            ]
        ],
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# CATEGORY REPORT
# =========================================================
st.divider()
st.subheader("🏷️ Category-wise Expense")


if expense_df.empty:

    st.info(
        "No expense transactions for this period."
    )

    category_report = pd.DataFrame(
        columns=[
            "Category",
            "Amount",
            "Percentage"
        ]
    )

else:

    category_report = (
        expense_df
        .groupby("category", as_index=False)["amount"]
        .sum()
        .sort_values(
            "amount",
            ascending=False
        )
    )

    if total_expense > 0:
        category_report["percentage"] = (
            category_report["amount"] /
            total_expense
        ) * 100
    else:
        category_report["percentage"] = 0

    category_report = category_report.rename(
        columns={
            "category": "Category",
            "amount": "Amount",
            "percentage": "Percentage"
        }
    )


    category_display = category_report.copy()

    category_display["Amount"] = (
        category_display["Amount"]
        .apply(money)
    )

    category_display["Percentage"] = (
        category_display["Percentage"]
        .apply(
            lambda x: f"{x:.1f}%"
        )
    )

    st.dataframe(
        category_display,
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# ACCOUNT REPORT
# =========================================================
st.divider()
st.subheader("🏦 Account-wise Report")


if report_df.empty:

    st.info(
        "No account transactions for this period."
    )

    account_report = pd.DataFrame()

else:

    account_report = (
        report_df
        .groupby(
            [
                "account",
                "type"
            ],
            as_index=False
        )["amount"]
        .sum()
    )

    account_report = (
        account_report
        .pivot(
            index="account",
            columns="type",
            values="amount"
        )
        .fillna(0)
        .reset_index()
    )

    if "income" not in account_report.columns:
        account_report["income"] = 0.0

    if "expense" not in account_report.columns:
        account_report["expense"] = 0.0

    account_report["balance"] = (
        account_report["income"]
        -
        account_report["expense"]
    )

    account_report = account_report.rename(
        columns={
            "account": "Account",
            "income": "Income",
            "expense": "Expense",
            "balance": "Balance"
        }
    )

    account_display = account_report.copy()

    for column in [
        "Income",
        "Expense",
        "Balance"
    ]:
        account_display[column] = (
            account_display[column]
            .apply(money)
        )

    st.dataframe(
        account_display[
            [
                "Account",
                "Income",
                "Expense",
                "Balance"
            ]
        ],
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# MONTHLY SUMMARY
# =========================================================
st.divider()
st.subheader("📅 Monthly Summary")


if report_df.empty:

    monthly_summary = pd.DataFrame(
        columns=[
            "Month",
            "Income",
            "Expense",
            "Balance"
        ]
    )

else:

    temp = report_df.copy()

    temp["Month"] = (
        temp["date"]
        .dt.to_period("M")
        .astype(str)
    )

    income_monthly = (
        temp[
            temp["type"] == "income"
        ]
        .groupby("Month")["amount"]
        .sum()
    )

    expense_monthly = (
        temp[
            temp["type"] == "expense"
        ]
        .groupby("Month")["amount"]
        .sum()
    )

    all_months = sorted(
        temp["Month"].unique()
    )

    monthly_rows = []

    for month in all_months:

        inc = float(
            income_monthly.get(
                month,
                0
            )
        )

        exp = float(
            expense_monthly.get(
                month,
                0
            )
        )

        monthly_rows.append({
            "Month": month,
            "Income": inc,
            "Expense": exp,
            "Balance": inc - exp
        })

    monthly_summary = pd.DataFrame(
        monthly_rows
    )


if not monthly_summary.empty:

    monthly_display = monthly_summary.copy()

    for column in [
        "Income",
        "Expense",
        "Balance"
    ]:

        monthly_display[column] = (
            monthly_display[column]
            .apply(money)
        )

    st.dataframe(
        monthly_display,
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# CHARTS
# =========================================================
st.divider()
st.subheader("📊 Charts")


if not report_df.empty:

    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:

        if not category_report.empty:

            chart_data = category_report.copy()

            chart_data["Amount"] = (
                chart_data["Amount"]
                .astype(float)
            )

            st.bar_chart(
                chart_data.set_index(
                    "Category"
                )["Amount"],
                use_container_width=True
            )

    with chart_col2:

        if not monthly_summary.empty:

            chart_monthly = monthly_summary.set_index(
                "Month"
            )[
                [
                    "Income",
                    "Expense"
                ]
            ]

            st.line_chart(
                chart_monthly,
                use_container_width=True
            )


# =========================================================
# FULL TRANSACTION REPORT
# =========================================================
st.divider()
st.subheader("📋 Transaction Details")


if report_df.empty:

    st.info(
        "No transactions available."
    )

else:

    transaction_display = report_df.copy()

    transaction_display["date"] = (
        transaction_display["date"]
        .dt.strftime("%d-%m-%Y")
    )

    transaction_display["type"] = (
        transaction_display["type"]
        .str.title()
    )

    transaction_display["amount"] = (
        transaction_display["amount"]
        .apply(money)
    )

    transaction_display = (
        transaction_display.rename(
            columns={
                "date": "Date",
                "type": "Type",
                "amount": "Amount",
                "category": "Category",
                "account": "Account",
                "note": "Note"
            }
        )
    )

    available_columns = [
        column
        for column in [
            "Date",
            "Type",
            "Amount",
            "Category",
            "Account",
            "Note"
        ]
        if column in transaction_display.columns
    ]

    st.dataframe(
        transaction_display[
            available_columns
        ],
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# EXCEL REPORT
# =========================================================
st.divider()
st.subheader("📥 Download Report")


def create_excel_report():

    wb = Workbook()

    ws = wb.active

    ws.title = "Yearly Report"


    # ========================================================
    # PREPARE DATA USING THE CURRENT REPORT FILTER
    # ========================================================
    # Keep the OLD Excel layout, but use data from Reports new.py.
    # The workbook remains ONE SHEET only.

    excel_income_categories = sorted(
        income_df["category"].dropna().astype(str).unique().tolist()
    ) if not income_df.empty else []

    excel_expense_categories = sorted(
        expense_df["category"].dropna().astype(str).unique().tolist()
    ) if not expense_df.empty else []

    def build_excel_category_table(source_df, categories):
        table = pd.DataFrame(
            0.0,
            index=categories,
            columns=months
        )

        if not source_df.empty:
            temp_source = source_df.copy()
            temp_source["Month"] = temp_source["date"].dt.month

            grouped = (
                temp_source
                .groupby(["category", "Month"])["amount"]
                .sum()
            )

            for category in categories:
                for month_number in range(1, 13):
                    try:
                        value = grouped.loc[
                            (category, month_number)
                        ]
                    except KeyError:
                        value = 0.0

                    table.loc[
                        category,
                        months[month_number - 1]
                    ] = float(value)

        table.index.name = "Category"
        table["Total"] = table.sum(axis=1)

        return table.reset_index()

    income_table = build_excel_category_table(
        income_df,
        excel_income_categories
    )

    expense_table = build_excel_category_table(
        expense_df,
        excel_expense_categories
    )

    monthly_income = []
    monthly_expense = []
    monthly_balance = []

    for month_number in range(1, 13):

        income_value = income_df.loc[
            income_df["date"].dt.month == month_number,
            "amount"
        ].sum()

        expense_value = expense_df.loc[
            expense_df["date"].dt.month == month_number,
            "amount"
        ].sum()

        monthly_income.append(float(income_value))
        monthly_expense.append(float(expense_value))
        monthly_balance.append(
            float(income_value - expense_value)
        )

    summary = pd.DataFrame({
        "Type": [
            "Income",
            "Expense",
            "Balance"
        ]
    })

    for i, month in enumerate(months):
        summary[month] = [
            monthly_income[i],
            monthly_expense[i],
            monthly_balance[i]
        ]

    summary["Year Total"] = [
        total_income,
        total_expense,
        balance
    ]

    spend_percent = expense_percent


    # ========================================================
    # COLORS
    # ========================================================

    title_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="5B9BD5"
    )

    section_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    total_fill = PatternFill(
        "solid",
        fgColor="E2F0D9"
    )


    # ========================================================
    # FONTS
    # ========================================================

    title_font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="FFFFFF"
    )

    section_font = Font(
        name="Calibri",
        size=13,
        bold=True
    )

    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF"
    )

    normal_font = Font(
        name="Calibri",
        size=11
    )

    bold_font = Font(
        name="Calibri",
        size=11,
        bold=True
    )


    # ========================================================
    # BORDER
    # ========================================================

    thin_side = Side(
        style="thin",
        color="B7B7B7"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )


    # ========================================================
    # TITLE
    # ========================================================

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=14
    )

    title_cell = ws.cell(
        1,
        1
    )

    title_cell.value = (
        f"MoneyMate - Report "
        f"{report_title}"
    )

    title_cell.fill = title_fill

    title_cell.font = title_font

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30


    # ========================================================
    # KPI
    # ========================================================

    kpis = [
        (
            "Total Income",
            total_income
        ),
        (
            "Total Expense",
            total_expense
        ),
        (
            "Balance",
            total_balance
        ),
        (
            "Spend %",
            spend_percent
        )
    ]


    for col, (label, value) in enumerate(
        kpis,
        start=1
    ):

        label_cell = ws.cell(
            3,
            col
        )

        label_cell.value = label

        label_cell.fill = section_fill

        label_cell.font = bold_font

        label_cell.border = border

        label_cell.alignment = Alignment(
            horizontal="center"
        )


        value_cell = ws.cell(
            4,
            col
        )

        value_cell.value = value

        value_cell.font = bold_font

        value_cell.border = border

        value_cell.alignment = Alignment(
            horizontal="center"
        )

        if label == "Spend %":

            value_cell.number_format = (
                '0.00"%"'
            )

        else:

            value_cell.number_format = (
                '#,##0.00'
            )


    # ========================================================
    # COMMON HEADERS
    # ========================================================

    table_headers = [
        "Category"
    ] + months + [
        "Total"
    ]


    # ========================================================
    # INCOME SECTION
    # ========================================================

    income_section_row = 7

    ws.merge_cells(
        start_row=income_section_row,
        start_column=1,
        end_row=income_section_row,
        end_column=14
    )

    cell = ws.cell(
        income_section_row,
        1
    )

    cell.value = "INCOME"

    cell.fill = section_fill

    cell.font = section_font


    income_header_row = (
        income_section_row + 1
    )


    for col, header in enumerate(
        table_headers,
        start=1
    ):

        cell = ws.cell(
            income_header_row,
            col
        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.border = border

        cell.alignment = Alignment(
            horizontal="center"
        )


    for row_index, row in income_table.iterrows():

        excel_row = (
            income_header_row +
            1 +
            row_index
        )

        for col_index, column in enumerate(
            table_headers,
            start=1
        ):

            cell = ws.cell(
                excel_row,
                col_index
            )

            cell.value = row[column]

            cell.font = normal_font

            cell.border = border

            if col_index > 1:

                cell.number_format = (
                    '#,##0.00'
                )


    income_total_row = (
        income_header_row +
        1 +
        len(income_table)
    )


    ws.cell(
        income_total_row,
        1
    ).value = "Total Income"


    for col in range(2, 15):

        column_letter = get_column_letter(
            col
        )

        cell = ws.cell(
            income_total_row,
            col
        )

        cell.value = (
            f"=SUM("
            f"{column_letter}"
            f"{income_header_row + 1}:"
            f"{column_letter}"
            f"{income_total_row - 1}"
            f")"
        )

        cell.number_format = (
            '#,##0.00'
        )


    for col in range(1, 15):

        cell = ws.cell(
            income_total_row,
            col
        )

        cell.fill = total_fill

        cell.font = bold_font

        cell.border = border


    # ========================================================
    # EXPENSE SECTION
    # ========================================================

    expense_section_row = (
        income_total_row + 3
    )

    ws.merge_cells(
        start_row=expense_section_row,
        start_column=1,
        end_row=expense_section_row,
        end_column=14
    )

    cell = ws.cell(
        expense_section_row,
        1
    )

    cell.value = "EXPENSE"

    cell.fill = section_fill

    cell.font = section_font


    expense_header_row = (
        expense_section_row + 1
    )


    for col, header in enumerate(
        table_headers,
        start=1
    ):

        cell = ws.cell(
            expense_header_row,
            col
        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.border = border

        cell.alignment = Alignment(
            horizontal="center"
        )


    for row_index, row in expense_table.iterrows():

        excel_row = (
            expense_header_row +
            1 +
            row_index
        )

        for col_index, column in enumerate(
            table_headers,
            start=1
        ):

            cell = ws.cell(
                excel_row,
                col_index
            )

            cell.value = row[column]

            cell.font = normal_font

            cell.border = border

            if col_index > 1:

                cell.number_format = (
                    '#,##0.00'
                )


    expense_total_row = (
        expense_header_row +
        1 +
        len(expense_table)
    )


    ws.cell(
        expense_total_row,
        1
    ).value = "Total Expense"


    for col in range(2, 15):

        column_letter = get_column_letter(
            col
        )

        cell = ws.cell(
            expense_total_row,
            col
        )

        cell.value = (
            f"=SUM("
            f"{column_letter}"
            f"{expense_header_row + 1}:"
            f"{column_letter}"
            f"{expense_total_row - 1}"
            f")"
        )

        cell.number_format = (
            '#,##0.00'
        )


    for col in range(1, 15):

        cell = ws.cell(
            expense_total_row,
            col
        )

        cell.fill = total_fill

        cell.font = bold_font

        cell.border = border


    # ========================================================
    # YEAR SUMMARY SECTION
    # ========================================================

    summary_section_row = (
        expense_total_row + 3
    )

    ws.merge_cells(
        start_row=summary_section_row,
        start_column=1,
        end_row=summary_section_row,
        end_column=14
    )

    cell = ws.cell(
        summary_section_row,
        1
    )

    cell.value = "YEAR SUMMARY"

    cell.fill = section_fill

    cell.font = section_font


    summary_header_row = (
        summary_section_row + 1
    )

    summary_headers = [
        "Type"
    ] + months + [
        "Year Total"
    ]


    for col, header in enumerate(
        summary_headers,
        start=1
    ):

        cell = ws.cell(
            summary_header_row,
            col
        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.border = border

        cell.alignment = Alignment(
            horizontal="center"
        )


    for row_index, row in summary.iterrows():

        excel_row = (
            summary_header_row +
            1 +
            row_index
        )

        for col_index, column in enumerate(
            summary_headers,
            start=1
        ):

            cell = ws.cell(
                excel_row,
                col_index
            )

            cell.value = row[column]

            cell.font = normal_font

            cell.border = border

            if col_index > 1:

                cell.number_format = (
                    '#,##0.00'
                )


    # ========================================================
    # CHART DATA
    # ========================================================

    chart_column = 17

    chart_header_row = 2


    ws.cell(
        chart_header_row,
        chart_column
    ).value = "Month"

    ws.cell(
        chart_header_row,
        chart_column + 1
    ).value = "Income"

    ws.cell(
        chart_header_row,
        chart_column + 2
    ).value = "Expense"


    for i, month in enumerate(
        months,
        start=1
    ):

        row = (
            chart_header_row + i
        )

        ws.cell(
            row,
            chart_column
        ).value = month

        ws.cell(
            row,
            chart_column + 1
        ).value = monthly_income[
            i - 1
        ]

        ws.cell(
            row,
            chart_column + 2
        ).value = monthly_expense[
            i - 1
        ]


    # ========================================================
    # BAR CHART
    # ========================================================

    chart = BarChart()

    chart.type = "col"

    chart.style = 10

    chart.title = (
        f"Monthly Income vs Expense "
        f"- {selected_year}"
    )

    chart.y_axis.title = "Amount"

    chart.x_axis.title = "Month"

    chart.height = 8

    chart.width = 16


    data_reference = Reference(
        ws,
        min_col=chart_column + 1,
        max_col=chart_column + 2,
        min_row=chart_header_row,
        max_row=chart_header_row + 12
    )


    category_reference = Reference(
        ws,
        min_col=chart_column,
        min_row=chart_header_row + 1,
        max_row=chart_header_row + 12
    )


    chart.add_data(
        data_reference,
        titles_from_data=True
    )

    chart.set_categories(
        category_reference
    )

    chart.legend.position = "b"


    chart_location_row = (
        summary_header_row + 6
    )


    ws.add_chart(
        chart,
        f"A{chart_location_row}"
    )


    # ========================================================
    # HIDE CHART DATA
    # ========================================================

    for col in range(
        chart_column,
        chart_column + 3
    ):

        ws.column_dimensions[
            get_column_letter(col)
        ].hidden = True


    # ========================================================
    # COLUMN WIDTHS
    # ========================================================

    ws.column_dimensions["A"].width = 30

    for col in range(2, 15):

        ws.column_dimensions[
            get_column_letter(col)
        ].width = 13


    # ========================================================
    # EXCEL SETTINGS
    # ========================================================

    # IMPORTANT:
    # No freeze panes.

    ws.freeze_panes = None

    ws.sheet_view.showGridLines = False


    # ========================================================
    # SAVE
    # ========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output


# =========================================================
# DOWNLOAD BUTTON
# =========================================================
excel_file = create_excel_report()

st.download_button(
    label="📥 Download Excel Report",
    data=excel_file,
    file_name=(
        f"MoneyMate_Report_"
        f"{report_title.replace(' ', '_')}.xlsx"
    ),
    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    use_container_width=True
)


# =========================================================
# REPORT INFORMATION
# =========================================================
st.divider()

info1, info2, info3 = st.columns(3)

with info1:
    st.metric(
        "Transactions",
        len(report_df)
    )

with info2:
    st.metric(
        "Categories",
        report_df["category"].nunique()
        if not report_df.empty
        else 0
    )

with info3:
    st.metric(
        "Accounts",
        report_df["account"].nunique()
        if not report_df.empty
        else 0
    )


# =========================================================
# FOOTER
# =========================================================
st.divider()

st.caption(
    f"MoneyMate • {report_title} • "
    f"Currency: {currency_name}"
)

