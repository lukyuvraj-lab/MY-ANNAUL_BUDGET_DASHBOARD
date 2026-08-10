import streamlit as st
import pandas as pd
from io import BytesIO

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="MoneyMate - Reports",
    page_icon="📊",
    layout="wide"
)

st.title("📊 MoneyMate - Yearly Report")


# ============================================================
# LOAD TRANSACTIONS FROM SUPABASE
# ============================================================

try:
    from utils.supabase_client import supabase

    response = (
        supabase
        .table("transactions")
        .select("*")
        .execute()
    )

    data = response.data or []

except Exception as e:
    st.error(f"Unable to load transactions: {e}")
    data = []


if not data:
    st.info("No transactions found.")
    st.stop()


df = pd.DataFrame(data)


# ============================================================
# CHECK REQUIRED COLUMNS
# ============================================================

required_columns = ["date", "amount", "type", "category"]

missing_columns = [
    col for col in required_columns
    if col not in df.columns
]

if missing_columns:
    st.error(
        "The transactions table is missing these columns: "
        + ", ".join(missing_columns)
    )
    st.stop()


# ============================================================
# PREPARE DATA
# ============================================================

df["date"] = pd.to_datetime(
    df["date"],
    errors="coerce"
)

df["amount"] = pd.to_numeric(
    df["amount"],
    errors="coerce"
).fillna(0)

df["type"] = (
    df["type"]
    .fillna("")
    .astype(str)
    .str.strip()
)

df["category"] = (
    df["category"]
    .fillna("Uncategorized")
    .astype(str)
    .str.strip()
)

df.loc[
    df["category"] == "",
    "category"
] = "Uncategorized"

df = df.dropna(
    subset=["date"]
).copy()


if df.empty:
    st.info("No valid transactions found.")
    st.stop()


# ============================================================
# YEAR SELECTOR
# ============================================================

available_years = sorted(
    df["date"].dt.year.unique(),
    reverse=True
)

selected_year = st.selectbox(
    "📅 Select Year",
    available_years
)

year_df = df[
    df["date"].dt.year == selected_year
].copy()


if year_df.empty:
    st.warning(
        "No transactions found for the selected year."
    )
    st.stop()


# ============================================================
# MONTH INFORMATION
# ============================================================

months = [
    "Jan", "Feb", "Mar", "Apr",
    "May", "Jun", "Jul", "Aug",
    "Sep", "Oct", "Nov", "Dec"
]

year_df["Month"] = year_df["date"].dt.month


# ============================================================
# INCOME / EXPENSE
# ============================================================

income_mask = (
    year_df["type"]
    .str.lower()
    == "income"
)

expense_mask = (
    year_df["type"]
    .str.lower()
    == "expense"
)

income_df = year_df[
    income_mask
].copy()

expense_df = year_df[
    expense_mask
].copy()


total_income = income_df["amount"].sum()
total_expense = expense_df["amount"].sum()
total_savings = total_income - total_expense

if total_income > 0:
    total_spend_percent = (
        total_expense / total_income
    ) * 100
else:
    total_spend_percent = 0


# ============================================================
# STREAMLIT SUMMARY
# ============================================================

st.subheader(
    f"📅 {selected_year} Overview"
)

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric(
        "💰 Total Income",
        f"₹{total_income:,.2f}"
    )

with col2:
    st.metric(
        "💸 Total Expense",
        f"₹{total_expense:,.2f}"
    )

with col3:
    st.metric(
        "🏦 Total Savings",
        f"₹{total_savings:,.2f}"
    )

with col4:
    st.metric(
        "📊 Total Spend %",
        f"{total_spend_percent:.2f}%"
    )

st.divider()


# ============================================================
# CREATE INCOME TABLE
# ============================================================

if income_df.empty:

    income_table = pd.DataFrame(
        columns=["Category"] + months + ["Total"]
    )

else:

    income_table = pd.pivot_table(
        income_df,
        index="category",
        columns="Month",
        values="amount",
        aggfunc="sum",
        fill_value=0
    )

    income_table = income_table.reindex(
        columns=range(1, 13),
        fill_value=0
    )

    income_table.columns = months

    income_table["Total"] = income_table.sum(axis=1)

    total_row = income_table.sum(axis=0)
    total_row.name = "Total"

    income_table = pd.concat(
        [income_table, total_row.to_frame().T]
    )

    income_table = (
        income_table
        .reset_index()
        .rename(columns={"category": "Category"})
    )


# ============================================================
# CREATE EXPENSE TABLE
# ============================================================

if expense_df.empty:

    expense_table = pd.DataFrame(
        columns=["Category"] + months + ["Total"]
    )

else:

    expense_table = pd.pivot_table(
        expense_df,
        index="category",
        columns="Month",
        values="amount",
        aggfunc="sum",
        fill_value=0
    )

    expense_table = expense_table.reindex(
        columns=range(1, 13),
        fill_value=0
    )

    expense_table.columns = months

    expense_table["Total"] = expense_table.sum(axis=1)

    total_row = expense_table.sum(axis=0)
    total_row.name = "Total"

    expense_table = pd.concat(
        [expense_table, total_row.to_frame().T]
    )

    expense_table = (
        expense_table
        .reset_index()
        .rename(columns={"category": "Category"})
    )


# ============================================================
# STREAMLIT INCOME
# ============================================================

st.subheader("💰 Income")

st.dataframe(
    income_table,
    use_container_width=True,
    hide_index=True
)


# ============================================================
# STREAMLIT EXPENSE
# ============================================================

st.subheader("💸 Expense")

st.dataframe(
    expense_table,
    use_container_width=True,
    hide_index=True
)


# ============================================================
# YEAR SUMMARY
# ============================================================

monthly_income = (
    income_df.groupby("Month")["amount"].sum()
    if not income_df.empty
    else pd.Series(dtype=float)
)

monthly_expense = (
    expense_df.groupby("Month")["amount"].sum()
    if not expense_df.empty
    else pd.Series(dtype=float)
)

summary = pd.DataFrame(
    index=["Income", "Expense", "Balance"],
    columns=months,
    dtype=float
)

for month_number in range(1, 13):

    month_name = months[month_number - 1]

    income_value = monthly_income.get(
        month_number,
        0
    )

    expense_value = monthly_expense.get(
        month_number,
        0
    )

    summary.loc["Income", month_name] = income_value
    summary.loc["Expense", month_name] = expense_value
    summary.loc["Balance", month_name] = (
        income_value - expense_value
    )

summary["Year Total"] = summary.sum(axis=1)

summary_display = (
    summary
    .reset_index()
    .rename(columns={"index": "Type"})
)


st.subheader("🏦 Year Summary")

st.dataframe(
    summary_display,
    use_container_width=True,
    hide_index=True
)


# ============================================================
# STREAMLIT CHARTS
# ============================================================

st.subheader("📊 Charts")

chart_data = summary_display[
    summary_display["Type"].isin(
        ["Income", "Expense"]
    )
].copy()

chart_data = chart_data.melt(
    id_vars="Type",
    value_vars=months,
    var_name="Month",
    value_name="Amount"
)

# Monthly Income vs Expense
fig_income_expense = None

try:
    import plotly.express as px

    fig_income_expense = px.bar(
        chart_data,
        x="Month",
        y="Amount",
        color="Type",
        barmode="group",
        title=(
            f"Monthly Income vs Expense - "
            f"{selected_year}"
        )
    )

    st.plotly_chart(
        fig_income_expense,
        use_container_width=True
    )

except Exception:
    st.warning(
        "Plotly chart could not be displayed."
    )


# Expense by Category
if not expense_df.empty:

    category_expense = (
        expense_df
        .groupby("category")["amount"]
        .sum()
        .reset_index()
        .sort_values(
            "amount",
            ascending=False
        )
    )

    try:
        fig_category = px.pie(
            category_expense,
            names="category",
            values="amount",
            title=(
                f"Expense by Category - "
                f"{selected_year}"
            )
        )

        st.plotly_chart(
            fig_category,
            use_container_width=True
        )

    except Exception:
        pass


# ============================================================
# EXCEL COLORS - MATCH SCREENSHOT STYLE
# ============================================================

DARK_BLUE = "4F64A8"
HEADER_BLUE = "D9E2F3"
LIGHT_BLUE = "E8EDF7"
VERY_LIGHT_BLUE = "EEF2F8"
LIGHT_GREY = "F2F2F2"
MEDIUM_GREY = "D9D9D9"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN = "70AD47"


# ============================================================
# EXCEL HELPER FUNCTIONS
# ============================================================

def set_cell(
    ws,
    cell,
    value=None,
    font=None,
    fill=None,
    border=None,
    alignment=None,
    number_format=None
):
    target = ws[cell]
    target.value = value

    if font:
        target.font = font

    if fill:
        target.fill = fill

    if border:
        target.border = border

    if alignment:
        target.alignment = alignment

    if number_format:
        target.number_format = number_format


def apply_row_fill(
    ws,
    row_number,
    start_col,
    end_col,
    fill
):
    for col in range(start_col, end_col + 1):
        ws.cell(
            row=row_number,
            column=col
        ).fill = fill


def style_table_header(
    ws,
    row_number,
    start_col,
    end_col
):
    for col in range(start_col, end_col + 1):

        cell = ws.cell(
            row=row_number,
            column=col
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_BLUE
        )

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=BLACK
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = Border(
            left=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            right=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            top=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            bottom=Side(
                style="thin",
                color=MEDIUM_GREY
            )
        )


def style_data_row(
    ws,
    row_number,
    start_col,
    end_col,
    grey=False
):
    fill = (
        PatternFill(
            "solid",
            fgColor=LIGHT_GREY
        )
        if grey
        else PatternFill(fill_type=None)
    )

    for col in range(start_col, end_col + 1):

        cell = ws.cell(
            row=row_number,
            column=col
        )

        cell.fill = fill

        cell.font = Font(
            name="Arial",
            size=9,
            color=BLACK
        )

        cell.alignment = Alignment(
            vertical="center",
            horizontal=(
                "left"
                if col == start_col
                else "right"
            )
        )

        cell.border = Border(
            left=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            right=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            top=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            bottom=Side(
                style="thin",
                color=MEDIUM_GREY
            )
        )

        if col > start_col:
            cell.number_format = '#,##0.00'


def style_total_row(
    ws,
    row_number,
    start_col,
    end_col
):
    for col in range(start_col, end_col + 1):

        cell = ws.cell(
            row=row_number,
            column=col
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_BLUE
        )

        cell.font = Font(
            name="Arial",
            size=9,
            bold=True,
            color=BLACK
        )

        cell.border = Border(
            top=Side(
                style="medium",
                color=DARK_BLUE
            ),
            bottom=Side(
                style="medium",
                color=DARK_BLUE
            ),
            left=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            right=Side(
                style="thin",
                color=MEDIUM_GREY
            )
        )

        cell.alignment = Alignment(
            vertical="center",
            horizontal=(
                "left"
                if col == start_col
                else "right"
            )
        )

        if col > start_col:
            cell.number_format = '#,##0.00'


# ============================================================
# CREATE ONE-SHEET EXCEL REPORT
# ============================================================

def create_excel_report():

    wb = Workbook()

    ws = wb.active
    ws.title = "Annual Budget"

    # --------------------------------------------------------
    # SHEET VIEW
    # --------------------------------------------------------

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B13"

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    widths = {
        "A": 25,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 4,
        "O": 14,
        "P": 14,
        "Q": 14,
        "R": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # --------------------------------------------------------
    # COLORS / FONTS
    # --------------------------------------------------------

    title_font = Font(
        name="Arial",
        size=22,
        bold=True,
        color=BLACK
    )

    section_font = Font(
        name="Arial",
        size=14,
        bold=True,
        color=BLACK
    )

    normal_font = Font(
        name="Arial",
        size=9,
        color=BLACK
    )

    bold_font = Font(
        name="Arial",
        size=9,
        bold=True,
        color=BLACK
    )

    white_bold_font = Font(
        name="Arial",
        size=10,
        bold=True,
        color=WHITE
    )

    thin_blue = Side(
        style="medium",
        color=DARK_BLUE
    )

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    ws.merge_cells("A1:M1")

    ws["A1"] = "ANNUAL BUDGET"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    ws.row_dimensions[1].height = 32

    for col in range(1, 14):
        ws.cell(
            row=2,
            column=col
        ).fill = PatternFill(
            "solid",
            fgColor=DARK_BLUE
        )

    ws.row_dimensions[2].height = 5

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    ws.merge_cells("A4:B4")

    ws["A4"] = "SUMMARY"
    ws["A4"].font = bold_font
    ws["A4"].fill = PatternFill(
        "solid",
        fgColor=HEADER_BLUE
    )
    ws["A4"].alignment = Alignment(
        horizontal="left"
    )

    summary_items = [
        ("A5", "Total monthly income", "B5", total_income),
        ("A6", "Total monthly expenses", "B6", total_expense),
        ("A8", "BALANCE", "B8", total_savings),
        (
            "A10",
            "PERCENTAGE OF INCOME SPENT",
            "B10",
            total_spend_percent
        )
    ]

    for label_cell, label, value_cell, value in summary_items:

        ws[label_cell] = label
        ws[label_cell].font = normal_font
        ws[label_cell].alignment = Alignment(
            horizontal="left"
        )

        ws[value_cell] = value
        ws[value_cell].font = bold_font
        ws[value_cell].alignment = Alignment(
            horizontal="right"
        )

        if label == "PERCENTAGE OF INCOME SPENT":
            ws[value_cell].number_format = "0.00%"
        else:
            ws[value_cell].number_format = '#,##0.00'

    # Summary backgrounds
    for row in [5, 6]:
        apply_row_fill(
            ws,
            row,
            1,
            2,
            PatternFill(
                "solid",
                fgColor=VERY_LIGHT_BLUE
            )
        )

    apply_row_fill(
        ws,
        8,
        1,
        2,
        PatternFill(
            "solid",
            fgColor=HEADER_BLUE
        )
    )

    apply_row_fill(
        ws,
        10,
        1,
        2,
        PatternFill(
            "solid",
            fgColor=HEADER_BLUE
        )
    )

    # Dashed-looking bottom borders
    for row in [5, 6]:
        for col in range(1, 3):
            ws.cell(
                row=row,
                column=col
            ).border = Border(
                bottom=Side(
                    style="dashed",
                    color=MEDIUM_GREY
                )
            )

    # --------------------------------------------------------
    # TOP SUMMARY CHART DATA
    # --------------------------------------------------------

    ws["Q3"] = "Type"
    ws["R3"] = "Amount"

    ws["Q4"] = "Income"
    ws["R4"] = total_income

    ws["Q5"] = "Expense"
    ws["R5"] = total_expense

    for cell in ["Q3", "R3"]:
        ws[cell].font = bold_font

    top_chart = BarChart()

    top_chart.type = "col"
    top_chart.style = 10
    top_chart.title = ""
    top_chart.y_axis.title = ""
    top_chart.x_axis.title = ""
    top_chart.height = 7
    top_chart.width = 10
    top_chart.legend = None

    data_ref = Reference(
        ws,
        min_col=18,
        min_row=3,
        max_row=5
    )

    cats_ref = Reference(
        ws,
        min_col=17,
        min_row=4,
        max_row=5
    )

    top_chart.add_data(
        data_ref,
        titles_from_data=True
    )

    top_chart.set_categories(
        cats_ref
    )

    top_chart.dataLabels = DataLabelList()
    top_chart.dataLabels.showVal = True

    try:
        top_chart.series[0].graphicalProperties.solidFill = DARK_BLUE
    except Exception:
        pass

    ws.add_chart(
        top_chart,
        "D4"
    )

    # Hide helper data
    ws.column_dimensions["Q"].hidden = True
    ws.column_dimensions["R"].hidden = True

    # --------------------------------------------------------
    # INCOME SECTION
    # --------------------------------------------------------

    income_title_row = 12
    income_header_row = 13
    income_start_row = 14

    ws.merge_cells(
        start_row=income_title_row,
        start_column=1,
        end_row=income_title_row,
        end_column=13
    )

    ws.cell(
        row=income_title_row,
        column=1
    ).value = "INCOME"

    ws.cell(
        row=income_title_row,
        column=1
    ).font = section_font

    headers = ["Item"] + months

    for col, header in enumerate(
        headers,
        start=1
    ):
        ws.cell(
            row=income_header_row,
            column=col
        ).value = header

    style_table_header(
        ws,
        income_header_row,
        1,
        13
    )

    # Green item header like screenshot
    ws.cell(
        row=income_header_row,
        column=1
    ).fill = PatternFill(
        "solid",
        fgColor=GREEN
    )

    # Income categories
    income_categories = []

    if not income_df.empty:
        income_categories = (
            income_df["category"]
            .drop_duplicates()
            .tolist()
        )

    current_row = income_start_row

    for index, category in enumerate(
        income_categories
    ):

        ws.cell(
            row=current_row,
            column=1
        ).value = category

        for month_number in range(1, 13):

            amount = income_df.loc[
                (
                    income_df["category"] == category
                )
                &
                (
                    income_df["Month"] == month_number
                ),
                "amount"
            ].sum()

            ws.cell(
                row=current_row,
                column=month_number + 1
            ).value = float(amount)

        style_data_row(
            ws,
            current_row,
            1,
            13,
            grey=(index % 2 == 1)
        )

        current_row += 1

    income_total_row = current_row

    ws.cell(
        row=income_total_row,
        column=1
    ).value = "Total"

    for month_number in range(1, 13):

        amount = monthly_income.get(
            month_number,
            0
        )

        ws.cell(
            row=income_total_row,
            column=month_number + 1
        ).value = float(amount)

    style_total_row(
        ws,
        income_total_row,
        1,
        13
    )

    # --------------------------------------------------------
    # EXPENSE SECTION
    # --------------------------------------------------------

    expense_title_row = income_total_row + 2
    expense_header_row = expense_title_row + 1
    expense_start_row = expense_header_row + 1

    ws.merge_cells(
        start_row=expense_title_row,
        start_column=1,
        end_row=expense_title_row,
        end_column=13
    )

    ws.cell(
        row=expense_title_row,
        column=1
    ).value = "EXPENSES"

    ws.cell(
        row=expense_title_row,
        column=1
    ).font = section_font

    for col, header in enumerate(
        headers,
        start=1
    ):
        ws.cell(
            row=expense_header_row,
            column=col
        ).value = header

    style_table_header(
        ws,
        expense_header_row,
        1,
        13
    )

    # Green item header
    ws.cell(
        row=expense_header_row,
        column=1
    ).fill = PatternFill(
        "solid",
        fgColor=GREEN
    )

    expense_categories = []

    if not expense_df.empty:
        expense_categories = (
            expense_df["category"]
            .drop_duplicates()
            .tolist()
        )

    current_row = expense_start_row

    for index, category in enumerate(
        expense_categories
    ):

        ws.cell(
            row=current_row,
            column=1
        ).value = category

        for month_number in range(1, 13):

            amount = expense_df.loc[
                (
                    expense_df["category"] == category
                )
                &
                (
                    expense_df["Month"] == month_number
                ),
                "amount"
            ].sum()

            ws.cell(
                row=current_row,
                column=month_number + 1
            ).value = float(amount)

        style_data_row(
            ws,
            current_row,
            1,
            13,
            grey=(index % 2 == 1)
        )

        current_row += 1

    expense_total_row = current_row

    ws.cell(
        row=expense_total_row,
        column=1
    ).value = "Total"

    for month_number in range(1, 13):

        amount = monthly_expense.get(
            month_number,
            0
        )

        ws.cell(
            row=expense_total_row,
            column=month_number + 1
        ).value = float(amount)

    style_total_row(
        ws,
        expense_total_row,
        1,
        13
    )

    # --------------------------------------------------------
    # BALANCE SECTION
    # --------------------------------------------------------

    balance_title_row = expense_total_row + 2

    ws.cell(
        row=balance_title_row,
        column=1
    ).value = "BALANCE"

    ws.cell(
        row=balance_title_row,
        column=1
    ).font = bold_font

    for month_number in range(1, 13):

        balance_value = (
            monthly_income.get(
                month_number,
                0
            )
            -
            monthly_expense.get(
                month_number,
                0
            )
        )

        ws.cell(
            row=balance_title_row,
            column=month_number + 1
        ).value = float(balance_value)

    for col in range(1, 14):

        cell = ws.cell(
            row=balance_title_row,
            column=col
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_BLUE
        )

        cell.font = Font(
            name="Arial",
            size=9,
            bold=True
        )

        cell.border = Border(
            top=Side(
                style="medium",
                color=DARK_BLUE
            ),
            bottom=Side(
                style="medium",
                color=DARK_BLUE
            ),
            left=Side(
                style="thin",
                color=MEDIUM_GREY
            ),
            right=Side(
                style="thin",
                color=MEDIUM_GREY
            )
        )

        cell.alignment = Alignment(
            horizontal=(
                "left"
                if col == 1
                else "right"
            )
        )

        if col > 1:
            cell.number_format = '#,##0.00'

    # --------------------------------------------------------
    # HIDDEN MONTHLY CHART DATA
    # --------------------------------------------------------

    helper_col = 20  # T

    ws.cell(
        row=1,
        column=helper_col
    ).value = "Month"

    ws.cell(
        row=1,
        column=helper_col + 1
    ).value = "Income"

    ws.cell(
        row=1,
        column=helper_col + 2
    ).value = "Expense"

    for i, month in enumerate(
        months,
        start=2
    ):

        ws.cell(
            row=i,
            column=helper_col
        ).value = month

        ws.cell(
            row=i,
            column=helper_col + 1
        ).value = float(
            monthly_income.get(
                i - 1,
                0
            )
        )

        ws.cell(
            row=i,
            column=helper_col + 2
        ).value = float(
            monthly_expense.get(
                i - 1,
                0
            )
        )

    # --------------------------------------------------------
    # HIDDEN CATEGORY CHART DATA
    # --------------------------------------------------------

    category_helper_col = 24  # X

    ws.cell(
        row=1,
        column=category_helper_col
    ).value = "Category"

    ws.cell(
        row=1,
        column=category_helper_col + 1
    ).value = "Expense"

    if not expense_df.empty:

        for i, row_data in enumerate(
            category_expense.itertuples(
                index=False
            ),
            start=2
        ):

            ws.cell(
                row=i,
                column=category_helper_col
            ).value = str(
                row_data[0]
            )

            ws.cell(
                row=i,
                column=category_helper_col + 1
            ).value = float(
                row_data[1]
            )

    # --------------------------------------------------------
    # MONTHLY INCOME VS EXPENSE CHART
    # --------------------------------------------------------

    monthly_chart = BarChart()

    monthly_chart.type = "col"
    monthly_chart.style = 10
    monthly_chart.title = ""
    monthly_chart.y_axis.title = ""
    monthly_chart.x_axis.title = ""
    monthly_chart.height = 10
    monthly_chart.width = 18

    monthly_data = Reference(
        ws,
        min_col=helper_col + 1,
        max_col=helper_col + 2,
        min_row=1,
        max_row=13
    )

    monthly_categories = Reference(
        ws,
        min_col=helper_col,
        min_row=2,
        max_row=13
    )

    monthly_chart.add_data(
        monthly_data,
        titles_from_data=True
    )

    monthly_chart.set_categories(
        monthly_categories
    )

    monthly_chart.dataLabels = DataLabelList()
    monthly_chart.dataLabels.showVal = True

    try:
        monthly_chart.series[0].graphicalProperties.solidFill = DARK_BLUE
        monthly_chart.series[1].graphicalProperties.solidFill = "B4BAD4"
    except Exception:
        pass

    # Put chart below balance
    monthly_chart_row = balance_title_row + 3

    ws.add_chart(
        monthly_chart,
        f"A{monthly_chart_row}"
    )

    # --------------------------------------------------------
    # EXPENSE CATEGORY PIE CHART
    # --------------------------------------------------------

    if not expense_df.empty:

        pie_chart = PieChart()

        pie_chart.title = ""
        pie_chart.height = 10
        pie_chart.width = 14

        pie_data = Reference(
            ws,
            min_col=category_helper_col + 1,
            min_row=1,
            max_row=1 + len(category_expense)
        )

        pie_categories = Reference(
            ws,
            min_col=category_helper_col,
            min_row=2,
            max_row=1 + len(category_expense)
        )

        pie_chart.add_data(
            pie_data,
            titles_from_data=True
        )

        pie_chart.set_categories(
            pie_categories
        )

        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = True

        ws.add_chart(
            pie_chart,
            f"N{monthly_chart_row}"
        )

    # --------------------------------------------------------
    # HIDE HELPER COLUMNS
    # --------------------------------------------------------

    for col_num in range(
        helper_col,
        helper_col + 3
    ):
        ws.column_dimensions[
            get_column_letter(col_num)
        ].hidden = True

    for col_num in range(
        category_helper_col,
        category_helper_col + 2
    ):
        ws.column_dimensions[
            get_column_letter(col_num)
        ].hidden = True

    # --------------------------------------------------------
    # ROW HEIGHTS
    # --------------------------------------------------------

    for row in range(
        1,
        monthly_chart_row + 5
    ):
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 18

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[12].height = 24
    ws.row_dimensions[expense_title_row].height = 24

    # --------------------------------------------------------
    # PAGE SETUP
    # --------------------------------------------------------

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_title_rows = "1:13"

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # --------------------------------------------------------
    # SAVE TO MEMORY
    # --------------------------------------------------------

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output


# ============================================================
# EXCEL DOWNLOAD
# ============================================================

st.divider()

st.subheader("📤 Excel Export")

st.write(
    "Download the Annual Budget report in the "
    "single-sheet Excel format."
)

try:

    excel_file = create_excel_report()

    st.download_button(
        label="📥 Download Annual Budget Excel",
        data=excel_file.getvalue(),
        file_name=(
            f"MoneyMate_Annual_Budget_"
            f"{selected_year}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

except Exception as e:

    st.error(
        f"Unable to create Excel report: {e}"
    )
